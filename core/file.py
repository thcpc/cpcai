import abc
import json

import os
import pathlib
import shutil
import zipfile

from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.workbook import Workbook




class File(abc.ABC):
    def __init__(self, *args):
        self._path = self._generate_path(*args)
        self._relative_path = self._generate_relative_path(*args)

    @classmethod
    def RootPath(cls, env: str = "prod") -> pathlib.Path:
        if env == "dev":
            path = pathlib.Path.cwd()
            while path.name != "edktranslation":
                path = path.parent
            return path
        else:
            path = pathlib.Path.cwd()
            return path.joinpath("_internal")

    @classmethod
    def db_file(cls):
        for root, dirs, files in os.walk(cls.RootPath()):
            for file in files:
                if file == "edktranslationls.sqlite.db":
                    return str(pathlib.Path(root).joinpath(file))
        raise "No DataBase File"

    @abc.abstractmethod
    def folder(self) -> pathlib.Path:
        ...

    @abc.abstractmethod
    def relative_folder(self) -> pathlib.Path:
        ...

    @classmethod
    def system_download_folder(cls):
        try:
            import winreg

            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Explorer\User Shell Folders"
            )
            downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
            path, _ = winreg.QueryValueEx(key, downloads_guid)
            downloads = os.path.expandvars(path)  # 展开环境变量
            return pathlib.Path(downloads)
        except:
            _folder = os.path.join(os.environ['USERPROFILE'], 'Downloads')
            if os.path.exists(_folder):
                return pathlib.Path(_folder)
            else:
                return pathlib.Path(pathlib.Path.home().joinpath("Downloads"))

    def _generate_path(self, *args):
        _path = self.folder()
        for arg in args:
            _path = _path.joinpath(str(arg))
        return _path

    def _generate_relative_path(self, *args):
        _path = self.relative_folder()
        if _path is not None:
            for arg in args:
                _path = _path.joinpath(str(arg))
            return _path
        return ""

    def _before(self):
        if not self.path.parent.exists():
            self.path.parent.mkdir(parents=True)

    def copy(self, src):
        self._before()
        shutil.copy(src, self.path)

    def copyfileobj(self, file):
        self._before()
        with open(self.path, "wb") as f:
            shutil.copyfileobj(file.stream, f)

    def write(self, mode, file_data):
        self._before()
        with open(self.path, mode) as f:
            f.write(file_data)

    @property
    def path(self):
        return self._path

    def relative_path_str(self):
        return str(self._relative_path)

    def __str__(self):
        return str(self.path)


class DocFile(File):
    def __init__(self, doc_id, file_name):
        super().__init__(doc_id, file_name)

    def folder(self) -> pathlib.Path:
        return self.RootPath().joinpath("Data").joinpath("documents")

    def relative_folder(self):
        return pathlib.Path("Data").joinpath("documents")

class TranslateFile(File):
    def __init__(self, doc_id, language_id ,file_name):
        super().__init__(doc_id, language_id, file_name)

    def folder(self) -> pathlib.Path:
        return self.RootPath().joinpath("Data").joinpath("documents")

    def relative_folder(self):
        return pathlib.Path("Data").joinpath("documents")


class LanguageFile(File):
    def __init__(self, doc_id, doc_language_id, file_name):
        super().__init__(doc_id, doc_language_id, file_name)

    def folder(self) -> pathlib.Path:
        return self.RootPath().joinpath("Data").joinpath("documents")

    def relative_folder(self):
        return pathlib.Path("Data").joinpath("documents")



class UncertainTextFile(File):
    def __init__(self, doc_id, doc_language_id, file_name):
        super().__init__(doc_id, doc_language_id, file_name)

    def write(self, mode, file_data):
        self._before()
        with open(self.path, mode, encoding='utf-8') as f:
            json.dump(file_data, f, ensure_ascii=False)

    def read(self):
        with open(self.path, 'r', encoding='utf-8') as f:
            return json.loads(f.read())

    def folder(self) -> pathlib.Path:
        return self.RootPath().joinpath("Data").joinpath("documents")

    def relative_folder(self):
        return pathlib.Path("Data").joinpath("documents")


class FinalTextFile(File):
    def __init__(self, doc_id, doc_language_id, file_name):
        super().__init__(doc_id, doc_language_id, file_name)

    def write(self, mode, file_data):
        self._before()
        with open(self.path, mode, encoding='utf-8') as f:
            json.dump(file_data, f, ensure_ascii=False)

    def read(self):
        with open(self.path, 'r', encoding='utf-8') as f:
            return json.loads(f.read())

    def folder(self) -> pathlib.Path:
        return self.RootPath().joinpath("Data").joinpath("documents")

    def relative_folder(self):
        return pathlib.Path("Data").joinpath("documents")


class DownLoadFile(File):
    def __init__(self, file_name):
        super().__init__(file_name)

    def folder(self) -> pathlib.Path:
        return File.system_download_folder()

    def relative_folder(self):
        ...

    def zip(self, files: list[dict]):
        with zipfile.ZipFile(self.path, "w", zipfile.ZIP_DEFLATED) as zipf:
            _have_write = []
            for file in files:
                for original_name, new_name in file.items():
                    # 检查文件是否存在
                    if os.path.exists(original_name) and new_name not in _have_write:
                        zipf.write(original_name, arcname=new_name)  # arcname 指定新文件名
                        _have_write.append(new_name)
                    else:
                        print(f"文件 {original_name} 不存在，跳过")

class DownLoadLanguageXlsxFile(File):

    def __init__(self, file_name):
        super().__init__(file_name)

    def folder(self) -> pathlib.Path:
        return File.system_download_folder()

    def relative_folder(self) -> pathlib.Path:
        pass


class DownLoadTRDocxFile(File):
    def __init__(self, file_name):
        super().__init__(file_name)

    def folder(self) -> pathlib.Path:
        return File.system_download_folder()

    def relative_folder(self) -> pathlib.Path:
        pass

class DownloadTRPdfFile(File):
    def __init__(self, file_name):
        super().__init__(file_name)

    def folder(self) -> pathlib.Path:
        return File.system_download_folder()

    def relative_folder(self) -> pathlib.Path:
        pass

    def write(self, mode, source_file):
        pythoncom.CoInitialize()
        convert(source_file, str(self.path))

class DownLoadFinalTextFile(File):
    def relative_folder(self) -> pathlib.Path:
        ...

    def __init__(self, file_name):
        super().__init__(file_name)

    def folder(self) -> pathlib.Path:
        return File.system_download_folder()

    def headers(self):
        return ["文档源语言", "excel映射文", "已译文", "是否需要AI翻译"]

    def body(self, key, value):
        _translate_word = "不翻译"
        if value.get("needAI") == Translate.NoNeedAI:
            _translate_word = "已翻译"
        elif value.get("needAI") == Translate.NeedAI:
            _translate_word = "AI翻译"
        return [value.get("text"),
                value.get("relText"),
                value.get("translateText"),
                _translate_word
                ]

    def download(self, data):
        wb = Workbook()
        ws = wb.active
        bold_font = Font(name='Calibri', size=12, bold=True)
        header_fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ws.append(self.headers())
        _cells_side = []

        for key, value in data.items():
            ws.append(self.body(key, value))
            _cells_side.append(value.get("needAI"))

        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.border = Border(left=Side(style='thick'),
                                         right=Side(style='thick'),
                                         top=Side(style='thick'),
                                         bottom=Side(style='thick'))
                    cell.font = bold_font
                else:
                    if cell.column == 4 and cell.row != 1:
                        _need_ai = _cells_side.pop(0)
                        if _need_ai == Translate.NeedAI:
                            cell.font = Font(name='Calibri', color="ae3ec9", bold=True)
                            cell.border = thin_border
                        elif _need_ai == Translate.NoNeedAI:
                            cell.font = Font(name='Calibri', color="0b956c", bold=True)
                            cell.border = thin_border
                        elif _need_ai == Translate.No:
                            cell.font = Font(name='Calibri', color="6b7280", bold=True)
                            cell.border = thin_border
                    else:
                        cell.font = Font(name='Calibri')
                        cell.border = thin_border
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_length + 20
        wb.save(str(self.path))
